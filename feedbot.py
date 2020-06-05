import ibapi
import time


from pyrogram import Client, Filters
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Logging init
import logging

logging.basicConfig(format='%(message)s')
log = logging.getLogger(__name__)

# App credentials
chat_id = -1001280269964
api_id = 1384288
api_hash = "4a7c0e08b8f2cbc08e1ac16d3a534079"

# Other global variables
filename = "/home/ktbsh/tmp/orders.xlsx"
parts = 0
emoji = 0
akce = 0
ticker = 0
expirace = 0
typ = 0
strike = 0
smer = 0
mnozstvi = 0
cena = 0

# Function definitions go here:
def transformExpiration(date):
    month = date[:3]
    if month == "Jan":
        month = "01"
    elif month == "Feb":
        month = "02"
    elif month == "Mar":
        month = "03"
    elif month == "Apr":
        month = "04"
    elif month == "May":
        month = "05"
    elif month == "Jun":
        month = "06"
    elif month == "Jul":
        month = "07"
    elif month == "Aug":
        month = "08"
    elif month == "Sep":
        month = "09"
    elif month == "Oct":
        month = "10"
    elif month == "Nov":
        month = "11"
    elif month == "Dec":
        month = "12"
    day, year = date[3:].split("'")
    if int(day) < 10:
        day = "0%d" % int(day)
    return "20%s%s%s" % (year, month, day)

def parcel(message):
    try:
        global parts
        global emoji 
        global akce 
        global ticker 
        global expirace 
        global typ 
        global strike 
        global smer 
        global mnozstvi 
        global cena 
        parts = message.split()
        emoji = parts[0]
        akce = parts[1]
        ticker = parts[2]
        expirace = transformExpiration(parts[3])
        typ, strike = parts[4].split("-")
        smer = parts[5]
        mnozstvi = parts[6]
        cena = parts [9]
        if (len(parts) == 10):
            return True
    except:
        print("Error: " + message)
        return False

def notify_error(message):
    app.send_message("me", "[FeedBot] Channel message failed to process: " + message )


def append_to_table(date):
    date = datetime.fromtimestamp(date).strftime('%Y-%m-%d %H:%M:%S')
    sheet.append((akce, ticker, expirace, typ, strike, smer, mnozstvi, cena, date, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    book.save(filename=filename)
    print("Sheet updated and saved")

# Instantiate an excel workbook (open existing or create new)
try:
    book = load_workbook(filename=filename)
    sheet = book.active
except FileNotFoundError:
    print("Spreadsheet not found, creating new.")
    book = Workbook()
    sheet = book.active
    sheet.append(("akce", "ticker", "expirace", "typ", "strike", "směr", "množství", "cena", "cas zpravy", "cas ulozeni"))
else:
    print("Existing worksheet found, using " + filename)
    


# Main
app = Client("my_account", api_id, api_hash)

@app.on_message(Filters.chat(chat_id) & Filters.text)
def echo(client, message):
    if (message.text.startswith("🍏")):
        if (parcel(message.text)):
            log.warning("[Instant Order]Full message: " + message.text)
            append_to_table(message.date)
            print(ibapi.runMe(ticker, expirace, typ, strike, smer, mnozstvi))
        else:
            notify_error(message.text)
    elif (message.text.startswith("🍇")):
        if (parcel(message.text)):
            log.warning("Full message: " + message.text)
            append_to_table(message.date)
            print(ibapi.runMe(ticker, expirace, typ, strike, smer, mnozstvi))
        else:
            notify_error(message.text)
    else:
        pass
app.run()  # Automatically start() and idle()
